"""Dump row labels + year headers from the CMA input/assessment sheets."""
import sys
import pathlib
import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

path = pathlib.Path(sys.argv[1] if len(sys.argv) > 1
                    else "CMA_Auto_Analytics_v9.xlsx")
sheets = sys.argv[2].split(",") if len(sys.argv) > 2 else [
    "0_Setup", "2_Operating_Statement", "3_Balance_Sheet", "4_DSCR",
    "5_Holding_Levels", "6_Working_Capital", "8_Performance_Indicators",
    "9_Cross_Checks",
]

wb = openpyxl.load_workbook(path, read_only=True, data_only=False)

for name in sheets:
    if name not in wb.sheetnames:
        print(f"!! sheet missing: {name}")
        continue
    ws = wb[name]
    print(f"\n########## {name} ({ws.max_row}x{ws.max_column}) ##########")
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 115),
                                         max_col=min(ws.max_column, 16)), start=1):
        cells = []
        for c in row:
            v = c.value
            if v is None:
                continue
            s = str(v)
            width = int(sys.argv[3]) if len(sys.argv) > 3 else 42
            s = s if len(s) <= width else s[:width - 3] + "..."
            cells.append(f"{c.coordinate}={s}")
        if cells:
            print(f"  r{r}: " + " | ".join(cells))

wb.close()
