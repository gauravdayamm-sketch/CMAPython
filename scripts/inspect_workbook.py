"""Quick structural peek at the CMA workbook: sheet names + top-left cells."""
import sys
import pathlib
import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

path = pathlib.Path(sys.argv[1] if len(sys.argv) > 1
                    else "CMA_Auto_Analytics_v9.xlsx")
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

print(f"Workbook: {path.name}")
print(f"Sheets ({len(wb.sheetnames)}):")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  - {name}  ({ws.max_row} rows x {ws.max_column} cols)")

# Sample the first sheet that looks like an operating statement
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name}: first 12 rows x 6 cols ===")
    for row in ws.iter_rows(min_row=1, max_row=12, max_col=6, values_only=True):
        cells = [str(c)[:28] if c is not None else "" for c in row]
        print("  | " + " | ".join(cells))
    break

wb.close()
