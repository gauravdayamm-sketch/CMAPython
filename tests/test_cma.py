"""Tests for the CMA workbook pipeline: ingestion, Form V engine, scrutiny.

A synthetic borrower is written into a copy of the real template once per
session, so the parser is exercised against the genuine sheet layout.
"""
import sqlite3
import pytest

from ingest.cma_workbook import (
    build_year_columns, parse_workbook, save_to_db,
)
from ingest.demo_fixture import BORROWER
from analytics.wc_assessment import assess_working_capital
from analytics.forecast import scrutinize_projections


# ── Year architecture ─────────────────────────────────────────────────────────

def test_year_columns_default_config():
    cols = build_year_columns(2022, 2029, 2025)
    assert len(cols) == 9                       # 8 years + dual column
    assert [c.statement_type for c in cols] == (
        ["audited", "audited", "estimated", "audited", "estimated"]
        + ["projected"] * 4
    )
    dual = cols[2]
    assert dual.is_dual and dual.fy_label == "2024-25"
    assert cols[3].fy_label == "2024-25"        # audited twin, same FY
    assert cols[0].fy_label == "2022-23"
    assert cols[-1].fy_label == "2029-30"
    assert cols[-1].fy_end == "2030-03-31"


def test_year_columns_invalid_config_rejected():
    with pytest.raises(ValueError):
        build_year_columns(2025, 2024, 2025)    # last before first
    with pytest.raises(ValueError):
        build_year_columns(2010, 2029, 2025)    # > 12 columns


# demo_wb / ingested fixtures are shared via conftest.py

# ── Parsing ───────────────────────────────────────────────────────────────────

def test_parse_demo_workbook_clean(demo_wb):
    path, expected = demo_wb
    data = parse_workbook(path)
    assert data.errors() == []
    assert data.borrower["name"] == BORROWER
    assert len(data.years) == 9
    assert data.proposals["fb_wc"] == {"existing": 180.0, "proposed": 220.0}
    assert data.thresholds["min_current_ratio"] == pytest.approx(1.17)


def test_derived_subtotals_match_independent_sums(demo_wb):
    """bs_tca recomputed here from raw inputs must equal the parser's value."""
    path, expected = demo_wb
    data = parse_workbook(path)
    y = data.year("2024-25", "audited")
    raw = expected["2024-25|audited"]

    tca_independent = (
        raw["bs_cash"] + raw["bs_govt_securities"] + raw["bs_fixed_deposits"]
        + raw["bs_receivables_domestic"] + raw["bs_receivables_export"]
        + raw["bs_bills_purchased"] + raw["bs_deferred_recv_1yr"]
        + raw["bs_rm_imported"] + raw["bs_rm_indigenous"] + raw["bs_sip"]
        + raw["bs_fg"] + raw["bs_spares_imported"] + raw["bs_spares_indigenous"]
        + raw["bs_adv_suppliers_rm"] + raw["bs_adv_tax"]
        + raw["bs_tufs"] + raw["bs_duty_receivables"]
        + raw["bs_gst_input"] + raw["bs_other_oca"]
    )
    assert y.lines["bs_tca"] == pytest.approx(tca_independent, abs=0.01)
    # Balance identity holds on every column
    for yc in data.years:
        assert yc.lines["bs_total_liabilities"] == pytest.approx(
            yc.lines["bs_total_assets"], abs=0.5)


def test_unbalanced_sheet_rejected(demo_wb, tmp_path, temp_db):
    import openpyxl
    path, _ = demo_wb
    broken = tmp_path / "broken.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["3_Balance_Sheet"]
    ws["C43"] = ws["C43"].value + 100.0        # equity +100 → Yr 1 unbalanced
    wb.save(broken)
    wb.close()

    data = parse_workbook(broken)
    assert any("does not balance" in e for e in data.errors())
    with pytest.raises(ValueError):
        save_to_db(data, db_path=temp_db)


def test_template_drift_detected(demo_wb, tmp_path):
    import openpyxl
    path, _ = demo_wb
    drifted = tmp_path / "drifted.xlsx"
    wb = openpyxl.load_workbook(path)
    wb["2_Operating_Statement"]["B13"] = "Something Else Entirely"
    wb.save(drifted)
    wb.close()

    data = parse_workbook(drifted)
    assert any("Template drift" in e for e in data.errors())


# ── Persistence ───────────────────────────────────────────────────────────────

def test_ingest_writes_statements_and_financials(ingested):
    data, bid, expected, db = ingested
    with sqlite3.connect(db) as conn:
        n_stmt = conn.execute(
            "SELECT COUNT(*) FROM cma_statement WHERE borrower_id=?", (bid,)
        ).fetchone()[0]
        fin = conn.execute("""
            SELECT fy_end, sales, total_assets FROM financials
            WHERE borrower_id=? ORDER BY fy_end
        """, (bid,)).fetchall()
    assert n_stmt == 9
    # Audited years only: 2022-23, 2023-24, 2024-25
    assert [r[0] for r in fin] == ["2023-03-31", "2024-03-31", "2025-03-31"]
    assert fin[2][1] == pytest.approx(1815.0, abs=0.5)   # net sales FY25


def test_reingest_replaces_not_duplicates(ingested):
    data, bid, expected, db = ingested
    save_to_db(data, db_path=db)                # ingest the same file again
    with sqlite3.connect(db) as conn:
        n_stmt = conn.execute(
            "SELECT COUNT(*) FROM cma_statement WHERE borrower_id=?", (bid,)
        ).fetchone()[0]
        n_dual = conn.execute(
            "SELECT COUNT(*) FROM cma_statement WHERE borrower_id=? AND is_dual=1",
            (bid,)).fetchone()[0]
    assert n_stmt == 9
    assert n_dual == 1


# ── Working capital assessment ────────────────────────────────────────────────

def test_form_v_math(ingested):
    """MPBF verified independently from the stored line values."""
    data, bid, expected, db = ingested
    a = assess_working_capital(BORROWER, db_path=db)
    assert a.error == ""

    y = a.year("2024-25", "audited")
    m = y["metrics"]
    lines = data.year("2024-25", "audited").lines

    wcg = lines["bs_tca"] - lines["bs_ocl_total"]
    nwc = lines["bs_tca"] - lines["bs_tcl"]
    assert m["wcg"] == pytest.approx(wcg, abs=0.01)
    assert m["mpbf_method1"] == pytest.approx(wcg - max(0.25 * wcg, nwc), abs=0.01)
    assert m["mpbf_method2"] == pytest.approx(
        wcg - max(0.25 * lines["bs_tca"], nwc), abs=0.01)
    # ABF = WCG − NWC must equal short-term bank borrowings (identity)
    assert m["abf"] == pytest.approx(lines["bs_stbb_total"], abs=0.01)


def test_dscr_math(ingested):
    data, bid, expected, db = ingested
    a = assess_working_capital(BORROWER, db_path=db)
    y = a.year("2024-25", "audited")
    lines = data.year("2024-25", "audited").lines

    nca    = lines["os_cash_accruals"] - lines["ds_accruals_used"]
    inst   = lines["ds_tl_inst_applicant"] + lines["ds_tl_inst_other"]
    int_tl = lines["os_interest_tl"]
    assert y["metrics"]["gross_dscr"] == pytest.approx(
        (nca + int_tl) / (inst + int_tl), abs=1e-3)
    assert a.dscr_avg_gross is not None


def test_healthy_borrower_no_flags_or_breaches(ingested):
    data, bid, expected, db = ingested
    a = assess_working_capital(BORROWER, db_path=db)
    assert a.red_flags == []
    assert all(v["status"] == "OK" for v in a.verdicts), \
        [v for v in a.verdicts if v["status"] != "OK"]


def test_dual_variance_flag_triggers(ingested):
    """If audited actuals deviate >10% from the borrower's own estimate,
    the projection-credibility flag must fire."""
    data, bid, expected, db = ingested
    with sqlite3.connect(db) as conn:
        conn.execute("""
            UPDATE cma_line SET value = 1450.0
            WHERE line_code = 'os_net_sales' AND stmt_id =
                (SELECT stmt_id FROM cma_statement
                 WHERE borrower_id=? AND is_dual=1)
        """, (bid,))
    a = assess_working_capital(BORROWER, db_path=db)
    assert any("projection credibility" in f for f in a.red_flags)


# ── Projection scrutiny ───────────────────────────────────────────────────────

def test_projections_at_historical_cagr_are_in_line(ingested):
    data, bid, expected, db = ingested
    scr = scrutinize_projections(BORROWER, db_path=db)
    assert scr["error"] == ""
    sales = scr["metrics"]["Net Sales"]
    assert sales["error"] == ""
    assert len(sales["projections"]) == 4
    assert all(p["verdict"] == "IN LINE" for p in sales["projections"])
    assert sales["hist_cagr"] == pytest.approx(0.10, abs=0.01)


def test_inflated_projection_flagged_aggressive(ingested):
    data, bid, expected, db = ingested
    with sqlite3.connect(db) as conn:
        conn.execute("""
            UPDATE cma_line SET value = value * 1.6
            WHERE line_code = 'os_net_sales' AND stmt_id =
                (SELECT stmt_id FROM cma_statement
                 WHERE borrower_id=? AND statement_type='projected'
                 ORDER BY col_index DESC LIMIT 1)
        """, (bid,))
    scr = scrutinize_projections(BORROWER, db_path=db)
    last = scr["metrics"]["Net Sales"]["projections"][-1]
    assert last["verdict"] == "AGGRESSIVE"
