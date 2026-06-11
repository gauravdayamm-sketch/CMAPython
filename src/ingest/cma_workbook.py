"""
CMA workbook ingestion — parses the SBI-LLMS-format CMA workbook
(0_Setup, 2_Operating_Statement, 3_Balance_Sheet, 4_DSCR) into SQLite.

Design:
  * The parser core works on plain value grids (dict[sheet] -> list[list]),
    so the same code serves openpyxl (file on disk), xlwings (live Excel
    book) and unit-test fixtures.
  * Only INPUT cells are read from the sheets. Every subtotal the workbook
    derives with formulas is recomputed in Python (single source of truth),
    then validated.
  * Year columns follow the 0_Setup architecture: audited years, then a
    dual Estimated/Audited pair for the prior FY, the current FY estimate,
    then projections. The dual estimate is kept (is_dual=1) — its variance
    against audited actuals measures the borrower's projection credibility.
"""
import sqlite3
import pathlib
from dataclasses import dataclass, field

ROOT = pathlib.Path(__file__).resolve().parents[2]
DB   = ROOT / "db" / "cma.sqlite"

SHEET_SETUP = "0_Setup"
SHEET_OS    = "2_Operating_Statement"
SHEET_BS    = "3_Balance_Sheet"
SHEET_DSCR  = "4_DSCR"

FIRST_YEAR_COL = 3          # column C (1-based) holds Yr 1
MAX_YEAR_COLS  = 12
BALANCE_TOL    = 0.5        # ₹ Cr — same tolerance the workbook uses


# ── Line registry: sheet row -> line code ─────────────────────────────────────

OS_INPUTS = {
    7:  "os_domestic_sales",     8:  "os_export_sales",
    9:  "os_other_op_income",    11: "os_excise",
    12: "os_other_deductions",   15: "os_rm_imported",
    16: "os_rm_indigenous",      18: "os_spares_imported",
    19: "os_spares_indigenous",  21: "os_power_fuel",
    22: "os_direct_labour",      23: "os_other_mfg",
    24: "os_depreciation",       25: "os_opening_sip",
    26: "os_closing_sip",        28: "os_opening_fg",
    29: "os_closing_fg",         32: "os_sga",
    34: "os_interest_wc",        35: "os_interest_tl",
    36: "os_other_finance",      40: "os_interest_unsecured",
    41: "os_income_investments", 42: "os_forex_gain",
    43: "os_profit_sale_fa",     44: "os_profit_sale_inv",
    46: "os_goodwill_writeoff",  47: "os_prelim_writeoff",
    48: "os_misc_writeoff",      49: "os_forex_loss",
    50: "os_loss_sale_fa",       55: "os_mat_credit",
    56: "os_tax_provision",      57: "os_prior_period",
    59: "os_dividend_paid",
}

BS_INPUTS = {
    7:  "bs_stbb_applicant",        8:  "bs_stbb_other_banks",
    9:  "bs_buyers_credit",         10: "bs_bills_discounted_memo",
    13: "bs_stb_others",            14: "bs_creditors_trade",
    15: "bs_creditors_lc",          16: "bs_customer_advances",
    17: "bs_provision_tax",         18: "bs_dividend_payable",
    19: "bs_statutory_dues",        20: "bs_tl_instalments_1yr",
    21: "bs_other_cl_provisions",   22: "bs_expenditure_provision",
    23: "bs_creditors_capital_goods", 24: "bs_creditors_expenses",
    25: "bs_others_cl",             29: "bs_debentures",
    30: "bs_pref_shares_red",       31: "bs_secured_tl",
    32: "bs_dpc",                   33: "bs_term_deposits",
    34: "bs_unsecured_promoters",   35: "bs_unsecured_others",
    36: "bs_capex_creditors_lt",    37: "bs_noncurrent_provisions",
    38: "bs_dtl",                   41: "bs_guarantees_memo",
    43: "bs_equity_capital",        44: "bs_pref_capital",
    45: "bs_general_reserve",       46: "bs_revaluation_reserve",
    47: "bs_pl_surplus",            48: "bs_other_reserves",
    49: "bs_security_premium",      50: "bs_warrants",
    51: "bs_share_application",     55: "bs_cash",
    56: "bs_govt_securities",       57: "bs_fixed_deposits",
    59: "bs_receivables_domestic",  60: "bs_receivables_export",
    61: "bs_bills_purchased",       62: "bs_deferred_recv_1yr",
    65: "bs_rm_imported",           66: "bs_rm_indigenous",
    68: "bs_sip",                   69: "bs_fg",
    70: "bs_spares_imported",       71: "bs_spares_indigenous",
    74: "bs_adv_suppliers_rm",      75: "bs_adv_tax",
    77: "bs_tufs",                  78: "bs_duty_receivables",
    79: "bs_gst_input",             80: "bs_other_oca",
    84: "bs_gross_block",           85: "bs_cwip",
    86: "bs_acc_depreciation",      89: "bs_inv_group_cos",
    90: "bs_other_investments_lt",  92: "bs_adv_capital_goods",
    93: "bs_deferred_recv_lt",      94: "bs_deposits",
    95: "bs_loans_group",           96: "bs_receivables_6m",
    97: "bs_ar_group",              98: "bs_receivables_directors",
    99: "bs_other_nca",             101: "bs_nonconsumable_spares",
    104: "bs_gross_intangibles",    105: "bs_amortisation",
    107: "bs_dta",
}

DSCR_INPUTS = {
    8:  "ds_accruals_used",
    11: "ds_tl_inst_applicant",
    12: "ds_tl_inst_other",
}

# Structural anchors: (sheet, row, substring expected in column B).
# A mismatch means the template layout drifted — refuse to parse blindly.
ANCHORS = [
    (SHEET_OS, 13, "net sales"),
    (SHEET_OS, 33, "operating profit before interest"),
    (SHEET_OS, 58, "net profit"),
    (SHEET_BS, 27, "total current liabilities"),
    (SHEET_BS, 52, "total net worth"),
    (SHEET_BS, 82, "total current assets"),
    (SHEET_BS, 108, "total assets"),
    (SHEET_DSCR, 17, "gross dscr"),
]

THRESHOLD_CELLS = {
    # key                      (row, col)  on 0_Setup (1-based)
    "min_current_ratio":        (26, 3),
    "min_dscr_any":             (27, 3),
    "min_dscr_avg":             (28, 3),
    "max_tol_tnw":              (29, 3),
    "max_debt_ebitda":          (30, 3),
    "min_icr":                  (31, 3),
    "sales_growth_floor":       (26, 6),
    "min_facr":                 (27, 6),
    "max_nwc_erosion":          (28, 6),
    "max_inv_days_increase":    (29, 6),
    "max_dso_increase":         (30, 6),
    "max_dpo_stretch":          (31, 6),
}


# ── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class YearColumn:
    col_index:      int          # 1-based position (Yr 1 = 1)
    start_year:     int          # 2024 for FY 2024-25
    fy_label:       str          # '2024-25'
    fy_end:         str          # '2025-03-31'
    statement_type: str          # audited | estimated | projected
    is_dual:        bool = False
    lines:          dict = field(default_factory=dict)


@dataclass
class CMAData:
    borrower:   dict = field(default_factory=dict)
    thresholds: dict = field(default_factory=dict)
    proposals:  dict = field(default_factory=dict)
    setup_meta: dict = field(default_factory=dict)   # project-finance overlay etc.
    years:      list = field(default_factory=list)
    issues:     list = field(default_factory=list)   # (level, message)
    source:     str  = ""

    def errors(self):
        return [m for lvl, m in self.issues if lvl == "ERROR"]

    def year(self, fy_label, statement_type=None):
        for y in self.years:
            if y.fy_label == fy_label and (
                    statement_type is None or y.statement_type == statement_type):
                return y
        return None


# ── Grid adapters ─────────────────────────────────────────────────────────────

def load_grids_openpyxl(path):
    """Read the four CMA sheets from a saved workbook into value grids."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    grids = {}
    for name in (SHEET_SETUP, SHEET_OS, SHEET_BS, SHEET_DSCR):
        if name not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet '{name}'")
        ws = wb[name]
        grids[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return grids


def load_grids_xlwings(book):
    """Read the four CMA sheets from a live xlwings Book (unsaved edits included)."""
    grids = {}
    for name in (SHEET_SETUP, SHEET_OS, SHEET_BS, SHEET_DSCR):
        sht = book.sheets[name]
        last = sht.used_range.last_cell
        vals = sht.range((1, 1), (last.row, last.column)).value
        if vals is None:
            vals = []
        elif not isinstance(vals, list):
            vals = [[vals]]
        elif vals and not isinstance(vals[0], list):
            vals = [vals]
        grids[name] = vals
    return grids


def _cell(grid, row, col):
    """1-based safe access into a value grid."""
    try:
        return grid[row - 1][col - 1]
    except IndexError:
        return None


def _num(value, default=0.0):
    """Coerce a workbook cell to float. Formula strings / text count as missing."""
    if value is None or isinstance(value, str):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


# ── Year architecture (mirrors 0_Setup section 5) ─────────────────────────────

def build_year_columns(first_year, last_year, current_year):
    """
    Reproduce the workbook's column scheme:
      cols 1..prior_pos-1   audited history
      col  prior_pos        prior FY as borrower's old ESTIMATE (dual)
      col  prior_pos+1      prior FY audited actuals (latest audited)
      col  prior_pos+2      current FY estimate
      cols prior_pos+3..    projections
    """
    if not (first_year < current_year <= last_year):
        raise ValueError(
            f"Year config invalid: first={first_year}, "
            f"current={current_year}, last={last_year}"
        )
    n_years    = last_year - first_year + 1
    total_cols = n_years + 1                    # +1 for the dual column
    if total_cols > MAX_YEAR_COLS:
        raise ValueError(f"{total_cols} columns needed; template caps at {MAX_YEAR_COLS}")
    prior_pos  = (current_year - 1) - first_year + 1

    def label(y):
        return f"{y}-{str(y + 1)[-2:]}"

    cols = []
    for k in range(1, total_cols + 1):
        if k < prior_pos:
            y, typ, dual = first_year + k - 1, "audited", False
        elif k == prior_pos:
            y, typ, dual = current_year - 1, "estimated", True
        elif k == prior_pos + 1:
            y, typ, dual = current_year - 1, "audited", False
        elif k == prior_pos + 2:
            y, typ, dual = current_year, "estimated", False
        else:
            y, typ, dual = current_year + (k - prior_pos - 2), "projected", False
        cols.append(YearColumn(
            col_index=k, start_year=y, fy_label=label(y),
            fy_end=f"{y + 1}-03-31", statement_type=typ, is_dual=dual,
        ))
    return cols


# ── Derived line computations (ported from the sheet formulas) ────────────────

def compute_os_derived(L):
    L["os_gross_sales"] = L["os_domestic_sales"] + L["os_export_sales"] + L["os_other_op_income"]
    L["os_net_sales"]   = L["os_gross_sales"] - L["os_excise"] - L["os_other_deductions"]
    L["os_total_rm"]     = L["os_rm_imported"] + L["os_rm_indigenous"]
    L["os_total_spares"] = L["os_spares_imported"] + L["os_spares_indigenous"]
    L["os_cost_of_production"] = (
        L["os_total_rm"] + L["os_total_spares"] + L["os_power_fuel"]
        + L["os_direct_labour"] + L["os_other_mfg"] + L["os_depreciation"]
        + L["os_opening_sip"] - L["os_closing_sip"]
    )
    L["os_cost_of_sales"] = L["os_cost_of_production"] + L["os_opening_fg"] - L["os_closing_fg"]
    L["os_opbi"] = L["os_net_sales"] - L["os_cost_of_sales"] - L["os_sga"]
    L["os_total_interest"] = L["os_interest_wc"] + L["os_interest_tl"] + L["os_other_finance"]
    L["os_opai"] = L["os_opbi"] - L["os_total_interest"]
    L["os_nonop_income"] = (
        L["os_interest_unsecured"] + L["os_income_investments"] + L["os_forex_gain"]
        + L["os_profit_sale_fa"] + L["os_profit_sale_inv"]
    )
    L["os_nonop_expense"] = (
        L["os_goodwill_writeoff"] + L["os_prelim_writeoff"] + L["os_misc_writeoff"]
        + L["os_forex_loss"] + L["os_loss_sale_fa"]
    )
    L["os_net_nonop"] = L["os_nonop_income"] - L["os_nonop_expense"]
    L["os_pbt"] = L["os_opai"] + L["os_net_nonop"]
    L["os_pat"] = L["os_pbt"] - L["os_tax_provision"] - L["os_prior_period"] + L["os_mat_credit"]
    L["os_retained"] = L["os_pat"] - L["os_dividend_paid"]
    L["os_ebitda"] = L["os_opbi"] + L["os_depreciation"]
    L["os_cash_accruals"] = L["os_pat"] + L["os_depreciation"]


def compute_bs_derived(L):
    L["bs_stbb_total"] = L["bs_stbb_applicant"] + L["bs_stbb_other_banks"] + L["bs_buyers_credit"]
    L["bs_ocl_total"] = sum(L[c] for c in (
        "bs_stb_others", "bs_creditors_trade", "bs_creditors_lc",
        "bs_customer_advances", "bs_provision_tax", "bs_dividend_payable",
        "bs_statutory_dues", "bs_tl_instalments_1yr", "bs_other_cl_provisions",
        "bs_expenditure_provision", "bs_creditors_capital_goods",
        "bs_creditors_expenses", "bs_others_cl",
    ))
    L["bs_tcl"] = L["bs_stbb_total"] + L["bs_ocl_total"]
    L["bs_ttl"] = sum(L[c] for c in (
        "bs_debentures", "bs_pref_shares_red", "bs_secured_tl", "bs_dpc",
        "bs_term_deposits", "bs_unsecured_promoters", "bs_unsecured_others",
        "bs_capex_creditors_lt", "bs_noncurrent_provisions", "bs_dtl",
    ))
    L["bs_tol"] = L["bs_tcl"] + L["bs_ttl"]
    L["bs_tnw"] = sum(L[c] for c in (
        "bs_equity_capital", "bs_pref_capital", "bs_general_reserve",
        "bs_revaluation_reserve", "bs_pl_surplus", "bs_other_reserves",
        "bs_security_premium", "bs_warrants", "bs_share_application",
    ))
    L["bs_total_liabilities"] = L["bs_tol"] + L["bs_tnw"]
    L["bs_cash_investments"]  = L["bs_cash"] + L["bs_govt_securities"] + L["bs_fixed_deposits"]
    L["bs_receivables_total"] = (
        L["bs_receivables_domestic"] + L["bs_receivables_export"]
        + L["bs_bills_purchased"] + L["bs_deferred_recv_1yr"]
    )
    L["bs_rm_total"]     = L["bs_rm_imported"] + L["bs_rm_indigenous"]
    L["bs_spares_total"] = L["bs_spares_imported"] + L["bs_spares_indigenous"]
    L["bs_inventory_total"] = L["bs_rm_total"] + L["bs_sip"] + L["bs_fg"] + L["bs_spares_total"]
    L["bs_oca_total"] = L["bs_tufs"] + L["bs_duty_receivables"] + L["bs_gst_input"] + L["bs_other_oca"]
    L["bs_tca"] = (
        L["bs_cash_investments"] + L["bs_receivables_total"] + L["bs_inventory_total"]
        + L["bs_adv_suppliers_rm"] + L["bs_adv_tax"] + L["bs_oca_total"]
    )
    L["bs_net_block"] = L["bs_gross_block"] + L["bs_cwip"] - L["bs_acc_depreciation"]
    L["bs_lt_investments"] = L["bs_inv_group_cos"] + L["bs_other_investments_lt"]
    L["bs_other_nca_total"] = sum(L[c] for c in (
        "bs_adv_capital_goods", "bs_deferred_recv_lt", "bs_deposits",
        "bs_loans_group", "bs_receivables_6m", "bs_ar_group",
        "bs_receivables_directors", "bs_other_nca",
    ))
    L["bs_nca_total"] = L["bs_lt_investments"] + L["bs_other_nca_total"] + L["bs_nonconsumable_spares"]
    L["bs_net_intangibles"] = L["bs_gross_intangibles"] - L["bs_amortisation"]
    L["bs_total_assets"] = (
        L["bs_tca"] + L["bs_net_block"] + L["bs_nca_total"]
        + L["bs_net_intangibles"] + L["bs_dta"]
    )
    L["bs_nwc"] = L["bs_tca"] - L["bs_tcl"]


# ── Parser core ───────────────────────────────────────────────────────────────

def _verify_anchors(grids, issues):
    ok = True
    for sheet, row, expected in ANCHORS:
        label = _cell(grids[sheet], row, 2)
        if not isinstance(label, str) or expected not in label.lower():
            issues.append((
                "ERROR",
                f"Template drift: {sheet} row {row} should contain "
                f"'{expected}', found {label!r}"
            ))
            ok = False
    return ok


def parse_grids(grids, source=""):
    """Parse value grids into CMAData. Pure function — no DB access."""
    data = CMAData(source=source)
    issues = data.issues

    if not _verify_anchors(grids, issues):
        return data

    setup = grids[SHEET_SETUP]
    data.borrower = {
        "name":            _cell(setup, 5, 3),
        "cif":             _cell(setup, 6, 3),
        "industry":        _cell(setup, 7, 3),
        "internal_rating": _cell(setup, 8, 3),
        "external_rating": _cell(setup, 9, 3),
    }
    if not data.borrower["name"] or not str(data.borrower["name"]).strip():
        issues.append(("ERROR", "Borrower name not set in 0_Setup C5"))
        return data
    data.borrower["name"] = str(data.borrower["name"]).strip()

    # Year configuration
    first   = _num(_cell(setup, 12, 3), default=None)
    last    = _num(_cell(setup, 13, 3), default=None)
    current = _num(_cell(setup, 14, 3), default=None)
    if None in (first, last, current):
        issues.append(("ERROR", "Year configuration incomplete in 0_Setup C12/C13/C14"))
        return data
    try:
        data.years = build_year_columns(int(first), int(last), int(current))
    except ValueError as e:
        issues.append(("ERROR", str(e)))
        return data

    # Thresholds + proposal limits
    for key, (r, c) in THRESHOLD_CELLS.items():
        v = _cell(setup, r, c)
        if not isinstance(v, str) and v is not None:
            data.thresholds[key] = float(v)
    for row, facility in ((20, "fb_wc"), (21, "fb_tl"), (22, "nfb")):
        existing = next((_num(_cell(setup, row, c), None) for c in (2, 3)
                         if _num(_cell(setup, row, c), None) is not None), None)
        proposed = next((_num(_cell(setup, row, c), None) for c in (4, 5)
                         if _num(_cell(setup, row, c), None) is not None), None)
        if existing is not None or proposed is not None:
            data.proposals[facility] = {"existing": existing, "proposed": proposed}

    # Project-finance overlay (0_Setup section 7) — used by regulatory red flags
    for key, row in (("project_finance", 80), ("project_phase", 81),
                     ("project_sector", 82), ("dcco_year", 83)):
        v = _cell(setup, row, 3)
        if v is not None and str(v).strip() not in ("", "—"):
            data.setup_meta[key] = str(v).strip()

    # Line items per year column
    for ycol in data.years:
        col = FIRST_YEAR_COL + ycol.col_index - 1
        L = ycol.lines
        for row, code in OS_INPUTS.items():
            L[code] = _num(_cell(grids[SHEET_OS], row, col))
        for row, code in BS_INPUTS.items():
            L[code] = _num(_cell(grids[SHEET_BS], row, col))
        for row, code in DSCR_INPUTS.items():
            L[code] = _num(_cell(grids[SHEET_DSCR], row, col))
        compute_os_derived(L)
        compute_bs_derived(L)

    _validate(data)
    return data


def _validate(data):
    """Internal consistency checks (mirrors 9_Cross_Checks where applicable)."""
    issues = data.issues
    for y in data.years:
        L = y.lines
        tag = f"{y.fy_label} [{y.statement_type}{'/dual' if y.is_dual else ''}]"

        imbalance = L["bs_total_liabilities"] - L["bs_total_assets"]
        if abs(imbalance) > BALANCE_TOL:
            issues.append((
                "ERROR",
                f"{tag}: balance sheet does not balance — "
                f"Liabilities−Assets = {imbalance:+.2f} Cr"
            ))

        if y.statement_type == "audited":
            if L["os_net_sales"] == 0:
                issues.append(("WARNING", f"{tag}: net sales is zero"))
            if L["bs_tca"] == 0:
                issues.append(("WARNING", f"{tag}: total current assets is zero"))
            if L["bs_tnw"] < 0:
                issues.append(("WARNING", f"{tag}: negative net worth "
                                          f"({L['bs_tnw']:.1f} Cr)"))
        for code in ("bs_inventory_total", "bs_receivables_total", "bs_tca"):
            if L[code] < 0:
                issues.append(("WARNING", f"{tag}: {code} is negative"))


def parse_workbook(path):
    """Parse a saved CMA workbook file into CMAData."""
    return parse_grids(load_grids_openpyxl(path), source=str(path))


# ── Persistence ───────────────────────────────────────────────────────────────

# Canonical mapping into the `financials` table consumed by the existing
# analytics modules (Ohlson, Beneish, forecasts). Actuals only.
FINANCIALS_MAP = {
    "total_assets":        "bs_total_assets",
    "total_liabilities":   "bs_tol",
    "current_assets":      "bs_tca",
    "current_liabilities": "bs_tcl",
    "working_capital":     "bs_nwc",
    "net_income":          "os_pat",
    "ebit":                "os_opbi",
    "ebitda":              "os_ebitda",
    "ffo":                 "os_cash_accruals",
    "depreciation":        "os_depreciation",
    "sga":                 "os_sga",
    "cogs":                "os_cost_of_sales",
    "sales":               "os_net_sales",
    "ppe":                 "bs_net_block",
    "long_term_debt":      "bs_ttl",
    "tnw":                 "bs_tnw",
    "tol":                 "bs_tol",
    "interest_cost":       "os_total_interest",
}


def save_to_db(data, db_path=None):
    """
    Persist parsed CMAData. Replaces any previous CMA import for the
    borrower. Returns the borrower_id.
    """
    if data.errors():
        raise ValueError(
            "Refusing to save: ingestion has errors:\n  "
            + "\n  ".join(data.errors())
        )
    db = db_path or DB

    with sqlite3.connect(db) as conn:
        conn.execute("""
            INSERT INTO borrower (name, cin, industry, rbi_sector)
            VALUES (?, ?, ?, NULL)
            ON CONFLICT(name) DO UPDATE SET
                cin = COALESCE(excluded.cin, cin),
                industry = COALESCE(excluded.industry, industry)
        """, (
            data.borrower["name"],
            data.borrower.get("cif"),
            data.borrower.get("industry"),
        ))
        bid = conn.execute(
            "SELECT borrower_id FROM borrower WHERE name = ?",
            (data.borrower["name"],)
        ).fetchone()[0]

        # Replace previous import wholesale
        conn.execute("""
            DELETE FROM cma_line WHERE stmt_id IN
                (SELECT stmt_id FROM cma_statement WHERE borrower_id = ?)
        """, (bid,))
        conn.execute("DELETE FROM cma_statement WHERE borrower_id = ?", (bid,))
        conn.execute("DELETE FROM cma_proposal WHERE borrower_id = ?", (bid,))
        conn.execute("DELETE FROM cma_threshold WHERE borrower_id = ?", (bid,))
        conn.execute("DELETE FROM cma_setup_meta WHERE borrower_id = ?", (bid,))

        for y in data.years:
            cur = conn.execute("""
                INSERT INTO cma_statement
                (borrower_id, fy_label, fy_end, statement_type, is_dual,
                 col_index, source_file)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (bid, y.fy_label, y.fy_end, y.statement_type,
                  1 if y.is_dual else 0, y.col_index, data.source))
            stmt_id = cur.lastrowid
            conn.executemany(
                "INSERT INTO cma_line (stmt_id, line_code, value) VALUES (?, ?, ?)",
                [(stmt_id, code, val) for code, val in y.lines.items()]
            )

            # Audited actuals also refresh the canonical financials table
            if y.statement_type == "audited":
                cols = {k: y.lines[v] for k, v in FINANCIALS_MAP.items()}
                cols["receivables"] = (y.lines["bs_receivables_domestic"]
                                       + y.lines["bs_receivables_export"])
                cols["securities"]  = (y.lines["bs_govt_securities"]
                                       + y.lines["bs_fixed_deposits"])
                names  = ", ".join(cols)
                marks  = ", ".join("?" for _ in cols)
                conn.execute(f"""
                    INSERT OR REPLACE INTO financials
                    (borrower_id, fy_end, audited, {names})
                    VALUES (?, ?, 1, {marks})
                """, (bid, y.fy_end, *cols.values()))

        for facility, lim in data.proposals.items():
            conn.execute("""
                INSERT INTO cma_proposal (borrower_id, facility, existing, proposed)
                VALUES (?, ?, ?, ?)
            """, (bid, facility, lim["existing"], lim["proposed"]))
        for key, value in data.thresholds.items():
            conn.execute("""
                INSERT INTO cma_threshold (borrower_id, key, value)
                VALUES (?, ?, ?)
            """, (bid, key, value))
        for key, value in data.setup_meta.items():
            conn.execute("""
                INSERT INTO cma_setup_meta (borrower_id, key, value)
                VALUES (?, ?, ?)
            """, (bid, key, value))

    return bid


def ingest_workbook(path, db_path=None):
    """Parse + persist in one call. Returns (CMAData, borrower_id or None)."""
    data = parse_workbook(path)
    if data.errors():
        return data, None
    bid = save_to_db(data, db_path=db_path)
    return data, bid
